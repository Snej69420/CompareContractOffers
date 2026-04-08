import sys
import copy
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout,
    QPushButton, QFileDialog, QTableView, QTabWidget, QHBoxLayout,
    QFormLayout, QGroupBox, QHeaderView, QLabel, QSplitter,
    QAbstractItemView, QMessageBox, QSizePolicy, QProgressDialog
)
from PySide6.QtCore import Qt, QAbstractTableModel, QThread, Signal
from PySide6.QtGui import QColor, QBrush

from src.DataHandler import DataHandler
from src.Contracts.compare import ContractProcessor
from src.UI.Compare import CompareTab


class AIWorker(QThread):
    """Runs the heavy AI comparisons in the background to prevent UI freezing."""
    finished = Signal(dict, dict)
    error = Signal(str)

    def __init__(self, processor, loaded_contracts):
        super().__init__()
        self.processor = processor
        self.loaded_contracts = loaded_contracts

    def run(self):
        try:
            self.processor.is_cancelled = False
            df_master = self.loaded_contracts[0]['data']
            master_dict = df_master.groupby('Categorie', sort=False)['Naam'].apply(lambda x: list(x.unique())).to_dict()

            target_contracts = {}
            for i in range(1, len(self.loaded_contracts)):
                if self.processor.is_cancelled:
                    raise InterruptedError("Cancelled")

                contract_name = f"Contract {i + 1}"
                df_target = self.loaded_contracts[i]['data']
                t_dict = df_target.groupby('Categorie', sort=False)['Naam'].apply(lambda x: list(x.unique())).to_dict()

                ai_results = self.processor.process(df_master, df_target)

                target_contracts[contract_name] = {
                    'data': t_dict,
                    'ai': ai_results
                }
            self.finished.emit(master_dict, target_contracts)

        except InterruptedError:
            self.error.emit("Cancelled")
        except Exception as e:
            self.error.emit(str(e))

    def cancel(self):
        self.processor.is_cancelled = True


class PandasModel(QAbstractTableModel):
    def __init__(self, data):
        super().__init__()
        self._data = data

    def rowCount(self, parent=None):
        return self._data.shape[0]

    def columnCount(self, parent=None):
        return self._data.shape[1]

    def data(self, index, role=Qt.ItemDataRole):
        if not index.isValid(): return None
        value = self._data.iloc[index.row(), index.column()]

        if role == Qt.DisplayRole:
            if isinstance(value, (int, float)):
                col_name = self._data.columns[index.column()]
                if "Prijs" in col_name: return f"€ {value:,.2f}"
                return f"{value:.2f}" if isinstance(value, float) else str(value)
            return str(value)

        if role == Qt.TextAlignmentRole:
            if isinstance(value, (int, float)):
                return Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter
            return Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter
        return None

    def headerData(self, section, orientation, role):
        if role == Qt.DisplayRole:
            if orientation == Qt.Horizontal: return str(self._data.columns[section])
            if orientation == Qt.Vertical: return str(section + 1)
        return None


class ContractWidget(QWidget):
    def __init__(self, contract_dict):
        super().__init__()
        self.metadata = contract_dict['metadata']
        self.df = contract_dict['data']
        self._setup_ui()
        self._merge_cells()

    def _setup_ui(self):
        main_layout = QVBoxLayout(self)
        meta_group = QGroupBox("Werf Informatie")
        meta_layout = QFormLayout()
        label_style = "font-weight: bold;"
        priority_keys = ["Werf Naam", "Werf Status", "Straatnaam en Nummer", "Postcode en Gemeente"]

        for key in priority_keys:
            if key in self.metadata:
                lbl_key = QLabel(key + ":")
                lbl_key.setStyleSheet(label_style)
                meta_layout.addRow(lbl_key, QLabel(str(self.metadata[key])))

        for key, val in self.metadata.items():
            if key not in priority_keys and key != "Bestandsnaam":
                lbl_key = QLabel(key + ":")
                lbl_key.setStyleSheet(label_style)
                meta_layout.addRow(lbl_key, QLabel(str(val)))

        meta_group.setLayout(meta_layout)

        self.table_view = QTableView()
        self.model = PandasModel(self.df)
        self.table_view.setModel(self.model)

        self.table_view.setAlternatingRowColors(True)
        self.table_view.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.table_view.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.table_view.horizontalHeader().setStretchLastSection(True)
        self.table_view.setSelectionBehavior(QTableView.SelectionBehavior.SelectRows)

        splitter = QSplitter(Qt.Orientation.Vertical)
        splitter.addWidget(meta_group)
        splitter.addWidget(self.table_view)
        splitter.setStretchFactor(1, 1)
        main_layout.addWidget(splitter)

    def _merge_cells(self):
        if 'Categorie' not in self.df.columns: return
        col_idx = self.df.columns.get_loc('Categorie')
        start_row = 0
        if self.df.empty: return
        current_val = self.df.iloc[0, col_idx]

        for i in range(1, len(self.df)):
            val = self.df.iloc[i, col_idx]
            if val != current_val:
                span_size = i - start_row
                if span_size > 1:
                    self.table_view.setSpan(start_row, col_idx, span_size, 1)
                current_val = val
                start_row = i

        span_size = len(self.df) - start_row
        if span_size > 1:
            self.table_view.setSpan(start_row, col_idx, span_size, 1)


class ContractApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Offerte Vergelijking - Native View")
        self.resize(1200, 800)
        self.data_handler = DataHandler()
        self.processor = ContractProcessor()
        self.loaded_contracts = []
        self._setup_ui()

    def _setup_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)

        btn_layout = QHBoxLayout()
        self.btn_load = QPushButton("📂 Laad Offertes")
        self.btn_load.setStyleSheet("padding: 8px; font-size: 14px; font-weight: bold;")
        self.btn_load.clicked.connect(self.load_files)

        self.btn_map = QPushButton("🤖 Vergelijk Offertes (AI)")
        self.btn_map.setStyleSheet("padding: 8px; font-size: 14px; font-weight: bold; color: #0056b3;")
        self.btn_map.clicked.connect(self.generate_mapping)
        self.btn_map.setEnabled(False)

        self.btn_export = QPushButton("💾 Exporteer Overzicht naar Excel")
        self.btn_export.setStyleSheet("padding: 10px; font-weight: bold; color: green;")
        self.btn_export.clicked.connect(self.export_data)
        self.btn_export.setEnabled(False)

        btn_layout.addWidget(self.btn_load)
        btn_layout.addWidget(self.btn_map)
        btn_layout.addWidget(self.btn_export)

        self.tabs = QTabWidget()
        self.tabs.setTabsClosable(True)
        self.tabs.tabCloseRequested.connect(self.close_tab)

        layout.addLayout(btn_layout)
        layout.addWidget(self.tabs)

    def load_files(self):
        files, _ = QFileDialog.getOpenFileNames(self, "Selecteer Offertes", "", "Excel Files (*.xlsx *.xls)")
        if not files: return
        contracts = self.data_handler.load_files(files)
        self.loaded_contracts.extend(contracts)

        for contract in contracts:
            page = ContractWidget(contract)
            tab_name = contract['metadata'].get('Bestandsnaam', 'Unknown')
            self.tabs.addTab(page, tab_name)
        self._update_button_states()

    def close_tab(self, index):
        tab_name = self.tabs.tabText(index)
        if tab_name != "⚙️ AI Mapping" and 0 <= index < len(self.loaded_contracts):
            del self.loaded_contracts[index]
        self.tabs.removeTab(index)
        self._update_button_states()

    def _update_button_states(self):
        self.btn_export.setEnabled(len(self.loaded_contracts) > 0)
        self.btn_map.setEnabled(len(self.loaded_contracts) >= 2)

    def generate_mapping(self):
        """Runs the semantic matcher via Thread and opens the Mapping Tab."""
        if len(self.loaded_contracts) < 2: return

        for i in range(self.tabs.count()):
            if self.tabs.tabText(i) == "⚙️ AI Mapping":
                self.tabs.removeTab(i)
                break

        # Show Loading Spinner Dialog
        self.progress_dialog = QProgressDialog("AI is berekeningen aan het maken...\nDit kan enkele seconden duren.",
                                               "Annuleer", 0, 0, self)
        self.progress_dialog.setWindowTitle("AI Mapping Maken")
        self.progress_dialog.setWindowModality(Qt.WindowModal)

        # Launch background worker
        self.worker = AIWorker(self.processor, self.loaded_contracts)
        self.worker.finished.connect(self.on_mapping_success)
        self.worker.error.connect(self.on_mapping_error)
        self.progress_dialog.canceled.connect(self.worker.cancel)

        self.worker.start()
        self.progress_dialog.show()

    def on_mapping_success(self, master_dict, target_contracts):
        self.progress_dialog.close()
        self.mapping_tab_widget = CompareTab(master_dict, target_contracts)
        tab_idx = self.tabs.addTab(self.mapping_tab_widget, "⚙️ AI Mapping")
        self.tabs.setCurrentIndex(tab_idx)

    def on_mapping_error(self, err_msg):
        self.progress_dialog.close()
        if err_msg != "Cancelled":
            QMessageBox.critical(self, "AI Fout", f"Mislukt:\n{err_msg}")

    def export_data(self):
        if not self.loaded_contracts: return
        file_path, _ = QFileDialog.getSaveFileName(self, "Save Comparison", "Offerte_Vergelijking.xlsx",
                                                   "Excel Files (*.xlsx)")
        if not file_path: return

        mapping_tab = None
        for i in range(self.tabs.count()):
            if self.tabs.tabText(i) == "⚙️ AI Mapping":
                mapping_tab = self.tabs.widget(i)
                break

        if not mapping_tab:
            QMessageBox.warning(self, "Fout", "Genereer eerst een AI Mapping voordat je exporteert.")
            return

        buckets = mapping_tab.extract_final_mapping()

        try:
            import pandas as pd
            import openpyxl
            from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
            from openpyxl.utils import get_column_letter
        except ImportError:
            QMessageBox.critical(self, "Fout",
                                 "De module 'openpyxl' ontbreekt. Installeer deze via 'pip install openpyxl'.")
            return

        try:
            lookups = {}
            columns_per_contract = {}

            # Extract raw lookup dictionaries for fast access
            for i, c in enumerate(self.loaded_contracts):
                c_name = f"Contract {i + 1}"
                df = c['data']
                cols = df.columns.tolist()
                columns_per_contract[c_name] = cols

                lookup = {}
                for _, row in df.iterrows():
                    key = (str(row['Categorie']).strip(), str(row['Naam']).strip())
                    lookup[key] = row.to_dict()
                lookups[c_name] = lookup

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Vergelijking"

            # Styling Setup
            fill_even = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

            align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
            align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

            border = Border(
                left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC')
            )
            header_font = Font(bold=True)

            contract_names = [f"Contract {i + 1}" for i in range(len(self.loaded_contracts))]
            current_col = 1
            col_offsets = {}

            # --- 1. WRITE HEADERS ---
            for c_name in contract_names:
                cols = columns_per_contract[c_name]
                col_offsets[c_name] = current_col

                # Merge Top Header for Contract Name
                ws.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=current_col + len(cols) - 1)
                cell = ws.cell(row=1, column=current_col, value=c_name)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = align_center

                for j in range(len(cols)):
                    ws.cell(row=1, column=current_col + j).border = border

                # Write Sub Headers (Column Names)
                for j, col_name in enumerate(cols):
                    c_cell = ws.cell(row=2, column=current_col + j, value=col_name)
                    c_cell.fill = header_fill
                    c_cell.font = header_font
                    c_cell.alignment = align_center
                    c_cell.border = border

                    # Dynamically set width
                    col_letter = get_column_letter(current_col + j)
                    if "Naam" in col_name or "Categorie" in col_name:
                        ws.column_dimensions[col_letter].width = 35
                    elif "Prijs" in col_name or "Totaal" in col_name:
                        ws.column_dimensions[col_letter].width = 15
                    else:
                        ws.column_dimensions[col_letter].width = 12

                current_col += len(cols)

            # --- 2. WRITE BUCKET DATA ---
            current_row = 3

            for bucket_idx, bucket in enumerate(buckets):
                max_rows = max([len(items) for items in bucket.values()] + [1])
                fill_color = fill_even if bucket_idx % 2 == 0 else fill_odd

                for c_name in contract_names:
                    items = bucket.get(c_name, [])
                    cols = columns_per_contract[c_name]
                    start_col = col_offsets[c_name]

                    # Write the physical rows first
                    for i in range(max_rows):
                        row_idx = current_row + i

                        if i < len(items):
                            cat, name = items[i]
                            row_data = lookups[c_name].get((cat, name), {})
                        else:
                            row_data = {}

                        for j, col_name in enumerate(cols):
                            val = row_data.get(col_name, "")
                            if pd.isna(val): val = ""

                            cell = ws.cell(row=row_idx, column=start_col + j, value=val)
                            cell.fill = fill_color
                            cell.border = border

                            if isinstance(val, (int, float)) and "Prijs" in col_name:
                                cell.number_format = '€ #,##0.00'

                            if "Naam" in col_name or "Categorie" in col_name:
                                cell.alignment = align_left
                            else:
                                cell.alignment = align_center

                    # Merge cells vertically if 1-to-Many exists
                    if len(items) == 1 and max_rows > 1:
                        for j in range(len(cols)):
                            ws.merge_cells(
                                start_row=current_row, start_column=start_col + j,
                                end_row=current_row + max_rows - 1, end_column=start_col + j
                            )

                current_row += max_rows

            ws.freeze_panes = 'A3'
            wb.save(file_path)
            QMessageBox.information(self, "Succes", f"Excel export succesvol opgeslagen:\n{file_path}")

        except Exception as e:
            QMessageBox.critical(self, "Fout", f"Exporteren is mislukt:\n{str(e)}")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = ContractApp()
    window.show()
    sys.exit(app.exec())