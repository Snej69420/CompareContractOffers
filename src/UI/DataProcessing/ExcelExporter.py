import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter


class ExcelExporter:
    NAME_WIDTH = 40
    NUMBER_WIDTH = 8
    UNIT_WIDTH = 4
    TOTAL_WIDTH = 10
    PCT_WIDTH = 7
    NOTES_WIDTH = 20

    def __init__(self, contract_names: dict[str, str]):
        self.contract_keys = list(contract_names.keys())
        self.contract_names = contract_names

        # Formatting styles
        self.yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        self.header_font = Font(bold=True)
        self.title_font = Font(bold=True, size=14)
        self.red_font = Font(color="FF0000")  # Kept ONLY for flagging parsing mismatches (Hoev/EH)

        # Alignments
        self.center_align = Alignment(horizontal="center", vertical="center")
        self.wrap_align = Alignment(vertical="center", wrap_text=True)
        self.valign_top = Alignment(vertical="top")

        # Borders
        thin = Side(border_style="thin", color="000000")
        self.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _safe_float(self, val) -> float:
        try:
            if isinstance(val, str): val = val.replace(',', '.')
            return float(val)
        except (ValueError, TypeError):
            return 0.0

    def _get_dominant(self, values: list):
        valid_vals = [v for v in values if v not in (None, "")]
        if not valid_vals: return None
        counts = Counter(valid_vals)
        most_common = counts.most_common(2)
        if len(most_common) == 1:
            return most_common[0][0]
        if most_common[0][1] > most_common[1][1]:
            return most_common[0][0]
        return None

    def export(self, norm_clusters: list[dict], norm_unmatched: list,
               filepath: str = "Tilia - Prjs VGL dakwerken.xlsx"):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Project Comparison"

        self._write_headers(ws)
        current_row = 3
        data_start_row = 3

        included_clusters = [c for c in norm_clusters if not c.get('is_excluded', False)]
        for cluster in included_clusters:
            current_row = self._write_cluster(ws, cluster, current_row)

        if norm_unmatched:
            ws.cell(row=current_row, column=1, value="ONGEKOPPELDE ITEMS").font = self.header_font
            current_row += 1
            for pseudo_cluster in norm_unmatched:
                current_row = self._write_cluster(ws, pseudo_cluster, current_row)

        data_end_row = current_row - 1

        excluded_clusters = [c for c in norm_clusters if c.get('is_excluded', False)]
        if excluded_clusters:
            current_row += 1
            ws.cell(row=current_row, column=1, value="ONGEVRAAGDE / BUITEN SCOPE").font = self.header_font
            current_row += 1
            for cluster in excluded_clusters:
                current_row = self._write_cluster(ws, cluster, current_row)

        # Overview at the very bottom
        self._write_overview(ws, current_row + 2, data_start_row, data_end_row)

        wb.save(filepath)

    def _write_overview(self, ws, overview_row: int, data_start: int, data_end: int):
        # 1. Setup row labels
        ws.cell(row=overview_row, column=1, value="OVERZICHT").font = self.title_font
        ws.cell(row=overview_row + 1, column=1, value="Totale Projectprijs:").font = self.header_font
        ws.cell(row=overview_row + 2, column=1, value="Verschil:").font = self.header_font

        # 2. Gather the exact left-most cells (col_offset) for the MIN function
        sum_cell_refs = []
        col_offset = 5
        for k in self.contract_keys:
            col_letter = get_column_letter(col_offset)  # Pointing to E, L, S
            sum_cell_refs.append(f"${col_letter}${overview_row + 1}")
            col_offset += 7

        # Build the absolute minimum reference formula (e.g., MIN($E$47,$L$47,$S$47))
        min_ref_formula = f"MIN({','.join(sum_cell_refs)})"

        col_offset = 5
        # Define a reusable left-alignment style
        left_align = Alignment(horizontal="left", vertical="center")

        for k in self.contract_keys:
            c_name = self.contract_names[k]

            sum_col_letter = get_column_letter(col_offset)
            tot_col_letter = get_column_letter(col_offset + 4)
            sum_cell_coord = f"{sum_col_letter}{overview_row + 1}"

            # Row 0: Name Header
            ws.merge_cells(start_row=overview_row, start_column=col_offset, end_row=overview_row,
                           end_column=col_offset + 6)
            name_cell = ws.cell(row=overview_row, column=col_offset, value=c_name.upper())
            name_cell.font = self.header_font
            name_cell.alignment = left_align

            # Row 1: The Total SUM
            sum_cell = ws.cell(row=overview_row + 1, column=col_offset)
            if data_end >= data_start:
                sum_cell.value = f"=SUM({tot_col_letter}{data_start}:{tot_col_letter}{data_end})"
            else:
                sum_cell.value = 0.0
            sum_cell.number_format = '#,##0.00'
            sum_cell.font = Font(bold=True)
            sum_cell.alignment = left_align  # Added Left Alignment
            ws.merge_cells(start_row=overview_row + 1, start_column=col_offset, end_row=overview_row + 1,
                           end_column=col_offset + 3)

            # Row 2: Percentage Difference
            pct_cell = ws.cell(row=overview_row + 2, column=col_offset)

            # Pure Dynamic Formula
            pct_cell.value = f"=IF({min_ref_formula}={sum_cell_coord}, 0, ({sum_cell_coord} / {min_ref_formula}) - 1)"
            pct_cell.number_format = '+0.0%;-0.0%;0.0%'
            pct_cell.font = Font(bold=True)
            pct_cell.alignment = left_align  # Changed to Left Alignment
            ws.merge_cells(start_row=overview_row + 2, start_column=col_offset, end_row=overview_row + 2,
                           end_column=col_offset + 6)

            col_offset += 7

    def _write_headers(self, ws):
        alg_header = ws.cell(row=1, column=1, value="ALGEMEEN")
        alg_header.font = self.header_font
        alg_header.alignment = self.center_align
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)

        for col in range(1, 4): ws.cell(row=1, column=col).border = self.border

        shared_headers = ["Naam", "Hoev", "EH"]
        for col_idx, header in enumerate(shared_headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = self.center_align

        ws.column_dimensions['A'].width = self.NAME_WIDTH
        ws.column_dimensions['B'].width = self.NUMBER_WIDTH
        ws.column_dimensions['C'].width = self.UNIT_WIDTH
        ws.column_dimensions['D'].width = 3

        col_offset = 5
        contractor_headers = ["Naam", "Hoev", "EH", "EP", "Tot", "%", "Notities"]

        for k in self.contract_keys:
            c_name = self.contract_names[k]

            header_cell = ws.cell(row=1, column=col_offset, value=c_name.upper())
            header_cell.font = self.header_font
            header_cell.alignment = self.center_align
            ws.merge_cells(start_row=1, start_column=col_offset, end_row=1, end_column=col_offset + 6)

            for col in range(col_offset, col_offset + 7): ws.cell(row=1, column=col).border = self.border

            for i, header in enumerate(contractor_headers):
                cell = ws.cell(row=2, column=col_offset + i, value=header)
                cell.font = self.header_font
                cell.border = self.border
                cell.alignment = self.center_align

            ws.column_dimensions[get_column_letter(col_offset)].width = self.NAME_WIDTH
            ws.column_dimensions[get_column_letter(col_offset + 1)].width = self.NUMBER_WIDTH
            ws.column_dimensions[get_column_letter(col_offset + 2)].width = self.UNIT_WIDTH
            ws.column_dimensions[get_column_letter(col_offset + 3)].width = self.NUMBER_WIDTH
            ws.column_dimensions[get_column_letter(col_offset + 4)].width = self.TOTAL_WIDTH
            ws.column_dimensions[get_column_letter(col_offset + 5)].width = self.PCT_WIDTH
            ws.column_dimensions[get_column_letter(col_offset + 6)].width = self.NOTES_WIDTH

            col_offset += 7

    def _write_cluster(self, ws, cluster: dict, start_row: int) -> int:
        lists = cluster.get('items', cluster)
        max_len = max(len(lists.get(k, [])) for k in self.contract_keys)
        if max_len == 0:
            return start_row

        end_row = start_row + max_len - 1

        contractor_totals = {}
        contractor_units = []
        original_units = []

        for k in self.contract_keys:
            items = lists.get(k, [])
            valid_items = [i for i in items if
                           not getattr(i, 'is_unmatched_blank', False) and not getattr(i, 'is_missing', False)]
            if valid_items:
                contractor_totals[k] = sum(self._safe_float(i.qty) for i in valid_items)
                for item in valid_items:
                    if item.unit:
                        contractor_units.append(str(item.unit).strip().lower())
                        original_units.append(item.unit)
            else:
                contractor_totals[k] = None

        dom_tot_qty = self._get_dominant([v for v in contractor_totals.values() if v is not None])
        dom_cluster_unit_lower = self._get_dominant(contractor_units)

        dom_cluster_unit = None
        if dom_cluster_unit_lower:
            for ou in original_units:
                if str(ou).strip().lower() == dom_cluster_unit_lower:
                    dom_cluster_unit = ou
                    break

        shared_qty_col_letter = "B"

        for i in range(max_len):
            current_row = start_row + i

            row_quantities, row_units = [], []
            for k in self.contract_keys:
                items = lists.get(k, [])
                if i < len(items):
                    item = items[i]
                    if not getattr(item, 'is_unmatched_blank', False) and not getattr(item, 'is_missing', False):
                        row_quantities.append(self._safe_float(item.qty))
                        row_units.append(str(item.unit).strip().lower() if item.unit else "")

            dom_row_qty = self._get_dominant(row_quantities)
            dom_row_unit_lower = self._get_dominant(row_units)

            dom_row_unit = None
            if dom_row_unit_lower:
                for k in self.contract_keys:
                    items = lists.get(k, [])
                    if i < len(items) and not getattr(items[i], 'is_missing', False):
                        u = items[i].unit
                        if u and str(u).strip().lower() == dom_row_unit_lower:
                            dom_row_unit = u
                            break

            cell_a = ws.cell(row=current_row, column=1, value="")
            cell_a.border = self.border
            cell_a.alignment = self.wrap_align

            cell_b = ws.cell(row=current_row, column=2)
            cell_b.border = self.border
            cell_b.alignment = self.center_align
            cell_b.number_format = '0.##'

            cell_c = ws.cell(row=current_row, column=3)
            cell_c.border = self.border
            cell_c.alignment = self.center_align

            if i == 0:
                if dom_tot_qty is not None:
                    cell_b.value = dom_tot_qty
                else:
                    cell_b.fill = self.yellow_fill

                if dom_cluster_unit is not None:
                    cell_c.value = dom_cluster_unit
                else:
                    cell_c.fill = self.yellow_fill

            col_offset = 5
            for k in self.contract_keys:
                items = lists.get(k, [])
                num_items = len(items)

                c_cells = [ws.cell(row=current_row, column=col_offset + j) for j in range(7)]
                for c in c_cells: c.border = self.border

                c_cells[0].alignment = self.wrap_align
                c_cells[1].alignment = self.center_align;
                c_cells[1].number_format = '0.##'
                c_cells[2].alignment = self.center_align
                c_cells[3].alignment = self.center_align;
                c_cells[3].number_format = '#,##0.00'
                c_cells[4].alignment = self.center_align;
                c_cells[4].number_format = '#,##0.00'
                c_cells[5].alignment = self.center_align
                c_cells[6].alignment = self.valign_top

                is_merged_contractor = (num_items <= 1)

                if (is_merged_contractor and i == 0) or (not is_merged_contractor and i < num_items):
                    item = items[0] if is_merged_contractor and num_items == 1 else (
                        items[i] if i < num_items else None)

                    if item:
                        if getattr(item, 'is_missing', False):
                            for c in c_cells: c.fill = self.yellow_fill

                        if not getattr(item, 'is_unmatched_blank', False):
                            qty = self._safe_float(item.qty)
                            ep = self._safe_float(item.raw_data.get('Eenheidsprijs', 0.0))

                            c_cells[0].value = item.name
                            c_cells[1].value = qty
                            c_cells[2].value = item.unit
                            c_cells[3].value = ep

                            if not getattr(item, 'is_missing', False):
                                # Red text ONLY for Mismatches against the Algemeen columns
                                if dom_tot_qty is not None and contractor_totals.get(k) is not None:
                                    if abs(contractor_totals[k] - dom_tot_qty) > 0.001:
                                        c_cells[1].font = self.red_font

                                if dom_cluster_unit is not None and item.unit:
                                    if str(item.unit).strip().lower() != dom_cluster_unit_lower:
                                        c_cells[2].font = self.red_font

                            c_qty_letter = get_column_letter(c_cells[1].column)
                            c_ep_letter = get_column_letter(c_cells[3].column)

                            if is_merged_contractor:
                                formula_tot = f"=IF(ISBLANK(${shared_qty_col_letter}${start_row}), {c_qty_letter}{current_row}, ${shared_qty_col_letter}${start_row}) * {c_ep_letter}{current_row}"
                            else:
                                formula_tot = f"={c_qty_letter}{current_row} * {c_ep_letter}{current_row}"

                            c_cells[4].value = formula_tot

                col_offset += 7

        if max_len > 1:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
            ws.merge_cells(start_row=start_row, start_column=3, end_row=end_row, end_column=3)

        col_offset = 5
        for k in self.contract_keys:
            items = lists.get(k, [])
            num_items = len(items)

            if max_len > 1 and num_items <= 1:
                for c_idx in [0, 1, 2, 3, 4, 6]:
                    ws.merge_cells(start_row=start_row, start_column=col_offset + c_idx, end_row=end_row,
                                   end_column=col_offset + c_idx)

            if max_len > 1:
                ws.merge_cells(start_row=start_row, start_column=col_offset + 5, end_row=end_row,
                               end_column=col_offset + 5)

            pct_cell = ws.cell(row=start_row, column=col_offset + 5)

            if sum(1 for i in items if
                   not getattr(i, 'is_unmatched_blank', False) and not getattr(i, 'is_missing', False)) > 0:
                tot_col_letter = get_column_letter(col_offset + 4)
                min_formula = f"MIN(" + ",".join(
                    [f"{get_column_letter(5 + j * 7 + 4)}{start_row}" for j in range(len(self.contract_keys))]) + ")"

                # Dynamic Row Percentage Formula
                pct_cell.value = f"=IF({min_formula}={tot_col_letter}{start_row}, 0, ({tot_col_letter}{start_row} / {min_formula}) - 1)"
                pct_cell.number_format = '+0.0%;-0.0%;0.0%'
            else:
                pct_cell.value = ""

            col_offset += 7

        return end_row + 1