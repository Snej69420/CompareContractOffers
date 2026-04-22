import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter


class ExcelExporter:
    def __init__(self, loaded_contracts):
        self.loaded_contracts = loaded_contracts

    def export(self, buckets, file_path):
        lookups = {}
        columns_per_contract = {}

        # 1. Build List-Based Lookups (Fixes the Duplicate Bug)
        for i, c in enumerate(self.loaded_contracts):
            c_name = f"Contract {i + 1}"
            df = c['data']
            cols = df.columns.tolist()
            columns_per_contract[c_name] = cols

            lookup = {}
            for _, row in df.iterrows():
                key = (str(row['Categorie']).strip(), str(row['Naam']).strip())
                if key not in lookup:
                    lookup[key] = []
                # Append to a list instead of overwriting!
                lookup[key].append(row.to_dict())
            lookups[c_name] = lookup

        # 2. Excel Setup
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Vergelijking"

        fill_even = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

        border = Border(
            left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC')
        )
        header_font = Font(bold=True)

        contract_names = [f"Contract {i + 1}" for i in range(len(self.loaded_contracts))]
        current_col = 1
        col_offsets = {}

        # 3. Write Headers
        for c_name in contract_names:
            cols = columns_per_contract[c_name]
            col_offsets[c_name] = current_col

            ws.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=current_col + len(cols) - 1)
            cell = ws.cell(row=1, column=current_col, value=c_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center

            for j in range(len(cols)):
                ws.cell(row=1, column=current_col + j).border = border

            for j, col_name in enumerate(cols):
                c_cell = ws.cell(row=2, column=current_col + j, value=col_name)
                c_cell.fill = header_fill
                c_cell.font = header_font
                c_cell.alignment = align_center
                c_cell.border = border

                col_letter = get_column_letter(current_col + j)
                if "Naam" in col_name or "Categorie" in col_name:
                    ws.column_dimensions[col_letter].width = 35
                elif "Prijs" in col_name or "Totaal" in col_name:
                    ws.column_dimensions[col_letter].width = 15
                else:
                    ws.column_dimensions[col_letter].width = 12

            current_col += len(cols)

        # 4. Write Data
        current_row = 3
        for bucket_idx, bucket in enumerate(buckets):
            max_rows = max([len(items) for items in bucket.values()] + [1])
            fill_color = fill_even if bucket_idx % 2 == 0 else fill_odd

            for c_name in contract_names:
                items = bucket.get(c_name, [])
                cols = columns_per_contract[c_name]
                start_col = col_offsets[c_name]

                for i in range(max_rows):
                    row_idx = current_row + i

                    if i < len(items):
                        cat, name = items[i]
                        key = (cat, name)

                        # POPOUT LOGIC: Pull the first available match and remove it from the list
                        if lookups[c_name].get(key):
                            row_data = lookups[c_name][key].pop(0)
                        else:
                            row_data = {}
                    else:
                        row_data = {}

                    for j, col_name in enumerate(cols):
                        val = row_data.get(col_name, "")
                        if pd.isna(val): val = ""

                        cell = ws.cell(row=row_idx, column=start_col + j, value=val)
                        cell.fill = fill_color
                        cell.border = border

                        if isinstance(val, (int, float)) and "Prijs" in col_name:
                            cell.number_format = '€ #,##0.00'

                        if "Naam" in col_name or "Categorie" in col_name:
                            cell.alignment = align_left
                        else:
                            cell.alignment = align_center

                # Merge cells vertically if 1-to-Many exists
                if len(items) == 1 and max_rows > 1:
                    for j in range(len(cols)):
                        ws.merge_cells(
                            start_row=current_row, start_column=start_col + j,
                            end_row=current_row + max_rows - 1, end_column=start_col + j
                        )

            current_row += max_rows

        ws.freeze_panes = 'A3'
        wb.save(file_path)